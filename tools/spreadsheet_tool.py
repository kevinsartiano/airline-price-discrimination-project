"""Tool to handle spreadsheet documents."""
import logging
import os
from collections import OrderedDict
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet


def generate_data(scraper):
    """Generate data to be exported."""
    return OrderedDict(
        {
            'os': scraper.user['os'],
            'browser': scraper.user['browser'],
            'ip_address': os.popen('curl -s ifconfig.me').read(),
            'search_date': str(datetime.now().strftime("%m-%d-%Y")),
            'search_time': str(datetime.now().strftime("%H:%M:%S")),
            'carrier': scraper.carrier,
            'origin': scraper.itinerary['origin'],
            'destination': scraper.itinerary['destination'],
            'fare_brand': scraper.itinerary['fare_brand'],
            'departure_date': to_date(scraper.itinerary['departure_date']),
            'departure_time': to_time(scraper.itinerary['departure_time']),
            'departure_flight': scraper.itinerary['departure_flight'],
            'departure_price': to_float(scraper.itinerary['departure_price']),
            'return_date': to_date(scraper.itinerary['return_date']),
            'return_time': to_time(scraper.itinerary['return_time']),
            'return_flight': scraper.itinerary['return_flight'],
            'return_price': to_float(scraper.itinerary['return_price']),
            'total_price': to_float(scraper.itinerary['total_price']),
            'control_price': to_float(scraper.itinerary['control_price']) - scraper.carrier_dcc,
            'dep_fare_basis': scraper.itinerary['dep_fare_basis'],
            'dep_control_fare_basis': scraper.itinerary['dep_control_fare_basis'],
            'ret_fare_basis': scraper.itinerary['ret_fare_basis'],
            'ret_control_fare_basis': scraper.itinerary['ret_control_fare_basis'],
            'seats_left': scraper.itinerary['seats_left']
        }
    )


def export_to_csv(scraper, dirname: str = 'output', basename: str = 'raw_data'):
    """
    Export to CSV file.

    Parameters:
        scraper (Scraper): scraper containing the data to export
        basename (str): file name. Example: /foo_1/foo_2/bar.xlsx -> bar
        dirname (str): directory path (defaults to 'output'). Example: /foo_1/foo_2/bar.xlsx -> /foo_1/foo_2
    """
    path = os.path.join(dirname, basename + '.xlsx')
    try:
        book = load_workbook(path)
    except FileNotFoundError:
        logging.warning('Spreadsheet missing. New one created')
        book = Workbook()
    sheet = book.active
    try:
        data = generate_data(scraper)
        if sheet.dimensions == 'A1:A1':
            sheet.append(list(data.keys()))
        sheet.append(list(data.values()))
    except KeyError as error:
        logging.warning(f'{scraper.identifier} | Missing {error}')
    adjust_column_width(sheet)
    sheet.freeze_panes = sheet['A2']
    sheet.auto_filter.ref = sheet.dimensions
    book.save(path)


def to_float(text: str) -> float or str:
    """Convert str to float if possible."""
    try:
        return round(float(text), 2)
    except ValueError:
        return text


def to_date(text: str) -> datetime or str:
    """Convert str to date if possible."""
    date_format = '%d/%m/%Y'
    try:
        return datetime.strptime(text, date_format).date()
    except ValueError:
        return text


def to_time(text: str) -> datetime or str:
    """Convert str to time if possible."""
    date_format = '%H:%M'
    try:
        return datetime.strptime(text, date_format).time()
    except ValueError:
        return text


def adjust_column_width(sheet: Worksheet):
    """Adjust the columns' width."""
    for i, column_cells in enumerate(sheet.columns, 1):
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 3
    for row in sheet.iter_rows(max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
